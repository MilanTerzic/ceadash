from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.utils.units import energy_to_mwh, ensure_datetime, mwh_to_mcm


@dataclass
class WorkbookData:
    gas_flows: pd.DataFrame
    historical_consumption: pd.DataFrame
    rs_to_hu: pd.DataFrame
    page_allocations: pd.DataFrame
    embedded_weather: pd.DataFrame
    metadata: dict[str, object]


class ExcelLoader:
    def __init__(self, workbook_path: str | Path, mcm_to_gwh: float) -> None:
        self.workbook_path = Path(workbook_path)
        self.mcm_to_gwh = mcm_to_gwh

    def load(self) -> WorkbookData:
        wb = load_workbook(self.workbook_path, read_only=True, data_only=True)
        gas_flows = self._load_flows(wb["flows"])
        historical = self._load_historical_consumption(wb["Historical SRB cons."])
        rs_to_hu = self._load_rs_to_hu(wb["RS to HU"])
        page_allocations = self._load_page(wb["page"])
        embedded_weather = self._load_embedded_weather(
            historical_consumption=historical,
            forecast_sheet=wb["Serbian Gas Cons. forecast"],
        )
        metadata = {
            "sheet_names": wb.sheetnames,
            "workbook_path": str(self.workbook_path),
        }
        return WorkbookData(
            gas_flows=gas_flows,
            historical_consumption=historical,
            rs_to_hu=rs_to_hu,
            page_allocations=page_allocations,
            embedded_weather=embedded_weather,
            metadata=metadata,
        )

    def _load_flows(self, sheet) -> pd.DataFrame:
        unit_row = [cell.value for cell in sheet[2]]
        records: list[dict[str, object]] = []
        point_columns = {
            3: "Kiskundorozsma (HU>RS)",
            4: "Kireevo (BG) / Zaychar (RS)",
            5: "Kiskundorozsma-2 (HU) / Horgos (RS)",
            6: "Kalotina",
        }
        for row in sheet.iter_rows(min_row=3, values_only=True):
            gas_day = row[1]
            for index, point_name in point_columns.items():
                raw_value = row[index - 1]
                if raw_value is None:
                    continue
                unit = unit_row[index - 1] or "kWh/d"
                value_mwh = energy_to_mwh(raw_value, unit)
                records.append(
                    {
                        "date": gas_day,
                        "point_name": point_name,
                        "country_from": "BG" if "BG" in point_name or "Kalotina" in point_name else "HU",
                        "country_to": "RS",
                        "direction": "inflow",
                        "value": raw_value,
                        "unit": unit,
                        "value_mwh": value_mwh,
                        "value_mcm": mwh_to_mcm(value_mwh, self.mcm_to_gwh),
                        "source": "excel",
                        "quality_flag": "actual",
                    }
                )
        frame = pd.DataFrame(records)
        if not frame.empty:
            frame["date"] = ensure_datetime(frame["date"]).dt.normalize()
        return frame

    def _load_historical_consumption(self, sheet) -> pd.DataFrame:
        records: list[dict[str, object]] = []
        for row in sheet.iter_rows(min_row=3, values_only=True):
            date = row[1]
            if date is None:
                continue
            records.append(
                {
                    "date": date,
                    "serbia_estimated_consumption_mcm": row[5],
                    "net_import_mcm": row[6],
                    "domestic_production_mcm": None,
                    "temperature_c": row[4],
                    "source": "excel",
                    "quality_flag": "calculated",
                }
            )
        frame = pd.DataFrame(records)
        if not frame.empty:
            frame["date"] = ensure_datetime(frame["date"]).dt.normalize()
        return frame

    def _load_rs_to_hu(self, sheet) -> pd.DataFrame:
        records: list[dict[str, object]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            value_mwh = energy_to_mwh(row[4], row[3])
            records.append(
                {
                    "date": row[0],
                    "point_name": row[2],
                    "country_from": "RS",
                    "country_to": "HU",
                    "direction": "outflow",
                    "value": row[4],
                    "unit": row[3],
                    "value_mwh": value_mwh,
                    "value_mcm": mwh_to_mcm(value_mwh, self.mcm_to_gwh),
                    "source": "excel",
                    "quality_flag": "actual",
                }
            )
        frame = pd.DataFrame(records)
        if not frame.empty:
            frame["date"] = ensure_datetime(frame["date"]).dt.normalize()
        return frame

    def _load_page(self, sheet) -> pd.DataFrame:
        rows = list(sheet.iter_rows(values_only=True))
        header = list(rows[0])
        frame = pd.DataFrame(rows[1:], columns=header)
        if frame.empty:
            return frame
        frame["periodFrom"] = ensure_datetime(frame["periodFrom"]).dt.normalize()
        frame["value_mwh"] = frame.apply(
            lambda row: energy_to_mwh(row.get("value"), row.get("unit")),
            axis=1,
        )
        frame["value_mcm"] = frame["value_mwh"].apply(
            lambda val: mwh_to_mcm(val, self.mcm_to_gwh)
        )
        frame["source"] = "excel_page"
        frame["quality_flag"] = "actual"
        return frame

    def _load_embedded_weather(self, historical_consumption: pd.DataFrame, forecast_sheet) -> pd.DataFrame:
        records: list[dict[str, object]] = []

        if not historical_consumption.empty and "temperature_c" in historical_consumption.columns:
            historical_weather = historical_consumption[["date", "temperature_c"]].dropna().copy()
            for row in historical_weather.itertuples(index=False):
                records.append(
                    {
                        "date": row.date,
                        "city": "Serbia",
                        "avg_temp_c": row.temperature_c,
                        "min_temp_c": row.temperature_c,
                        "max_temp_c": row.temperature_c,
                        "source": "excel_historical",
                        "quality_flag": "actual",
                    }
                )

        for row in forecast_sheet.iter_rows(min_row=7, values_only=True):
            date_value = row[1] if len(row) > 1 else None
            temp_value = row[4] if len(row) > 4 else None
            avg_temp_value = row[5] if len(row) > 5 else None
            effective_temp = avg_temp_value if avg_temp_value is not None else temp_value
            if date_value is None or effective_temp is None:
                continue
            records.append(
                {
                    "date": date_value,
                    "city": "Serbia",
                    "avg_temp_c": effective_temp,
                    "min_temp_c": effective_temp,
                    "max_temp_c": effective_temp,
                    "source": "excel_forecast_sheet",
                    "quality_flag": "actual",
                }
            )

        frame = pd.DataFrame(records)
        if frame.empty:
            return frame
        frame["date"] = ensure_datetime(frame["date"]).dt.normalize()
        frame = frame.dropna(subset=["date", "avg_temp_c"])
        frame = frame.sort_values("date").drop_duplicates(subset=["date", "city"], keep="last")
        return frame
